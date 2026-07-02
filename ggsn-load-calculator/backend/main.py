import os
import shutil
import sqlite3
import json
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Dict, Any, List, Optional
import openpyxl
import asyncio
import datetime

app = FastAPI(title="GGSN Load Calculator API")

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

DB_FILE = os.path.join(os.path.dirname(__file__), "ggsn_persistent_store.db")

def init_db():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Store settings / headers
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS system_settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    
    # Store the 3 sub-tables
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS table_license (
            node TEXT PRIMARY KEY,
            vendor TEXT,
            area TEXT,
            lic_bear REAL,
            lic_throughput REAL,
            lic_bear_uctt REAL,
            lic_throughput_vhkt REAL
        )
    """)
    # Migration: add area column if it does not yet exist (safe on existing DB)
    try:
        cursor.execute("ALTER TABLE table_license ADD COLUMN area TEXT DEFAULT ''")
    except Exception:
        pass  # Column already exists
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS table_current (
            node TEXT PRIMARY KEY,
            vendor TEXT,
            bear_su_dung REAL,
            throughput REAL,
            bear_total_su_dung REAL,
            bear_ims REAL,
            ipv4_internet REAL,
            ipv4_ims REAL
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS table_weight (
            node TEXT PRIMARY KEY,
            vendor TEXT,
            weight REAL,
            new_weight REAL,
            on_off INTEGER
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS table_ims_routing (
            node TEXT PRIMARY KEY,
            vendor TEXT,
            ims_site TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS table_hw_site (
            node TEXT PRIMARY KEY,
            vendor TEXT,
            hw_nfvi_site TEXT
        )
    """)
    
    # Check if empty, insert default initial nodes
    cursor.execute("SELECT COUNT(*) FROM table_license")
    if cursor.fetchone()[0] == 0:
        default_nodes = [
            ("GGPD04", "Huawei", "KV1", 2500000, 100000, 2750000, 110000),
            ("GGPD05", "Huawei", "KV1", 2500000, 100000, 2750000, 110000),
            ("GGPD06", "Huawei", "KV1", 3500000, 120000, 3850000, 130000),
            ("GGPD07", "Huawei", "KV1", 3500000, 120000, 3850000, 130000),
            ("GGHL04", "ZTE",   "KV2", 1090909, 45454, 1200000, 50000),
            ("GGHL05", "ZTE",   "KV2", 1090909, 45454, 1200000, 50000),
            ("GGHL06", "ZTE",   "KV2", 1090909, 45454, 1200000, 50000),
            ("GGHL07", "ZTE",   "KV2", 1090909, 45454, 1200000, 50000),
            ("GGHL11", "Ericsson", "KV3", 2500000, 100000, 2750000, 110000),
            ("GGHL12", "Ericsson", "KV3", 2500000, 100000, 2750000, 110000),
        ]
        cursor.executemany("INSERT INTO table_license VALUES (?, ?, ?, ?, ?, ?, ?)", default_nodes)
        
        default_current = [
            ("GGPD04", "Huawei", 2333754, 88777.52, 2333754, 1166877, 2000000, 350000),
            ("GGPD05", "Huawei", 2306341, 88791.13, 2306341, 1153170, 2000000, 340000),
            ("GGPD06", "Huawei", 3075825, 116824.07, 3075825, 1537912, 3000000, 480000),
            ("GGPD07", "Huawei", 3048884, 116981.69, 3048884, 1524442, 3000000, 470000),
            ("GGHL04", "ZTE", 1030469, 40625.70, 1030469, 515234, 1000000, 200000),
            ("GGHL05", "ZTE", 1037096, 40537.58, 1037096, 518548, 1000000, 180000),
            ("GGHL06", "ZTE", 299166, 3101.05, 299166, 149583, 1000000, 190000),
            ("GGHL07", "ZTE", 1118598, 42880.41, 1118598, 559299, 1000000, 200000),
            ("GGHL11", "Ericsson", 2490029, 89261.57, 2490029, 1245014, 2000000, 360000),
            ("GGHL12", "Ericsson", 2474712, 87932.89, 2474712, 1237356, 2000000, 340000),
        ]
        cursor.executemany("INSERT INTO table_current VALUES (?, ?, ?, ?, ?, ?, ?, ?)", default_current)
        
        default_weight = [
            ("GGPD04", "Huawei", 90, 90, 1),
            ("GGPD05", "Huawei", 90, 90, 1),
            ("GGPD06", "Huawei", 120, 120, 1),
            ("GGPD07", "Huawei", 120, 120, 1),
            ("GGHL04", "ZTE", 60, 60, 1),
            ("GGHL05", "ZTE", 60, 60, 1),
            ("GGHL06", "ZTE", 60, 60, 1),
            ("GGHL07", "ZTE", 60, 60, 1),
            ("GGHL11", "Ericsson", 90, 90, 1),
            ("GGHL12", "Ericsson", 90, 90, 1),
        ]
        cursor.executemany("INSERT INTO table_weight VALUES (?, ?, ?, ?, ?)", default_weight)
        
    cursor.execute("SELECT COUNT(*) FROM table_ims_routing")
    if cursor.fetchone()[0] == 0:
        default_ims = [
            ("GGPD04", "Huawei", "IMS_HN"),
            ("GGPD05", "Huawei", "IMS_HN"),
            ("GGPD06", "Huawei", "IMS_HN"),
            ("GGPD07", "Huawei", "IMS_HN"),
            ("GGHL04", "ZTE", "IMS_HN"),
            ("GGHL05", "ZTE", "IMS_HN"),
            ("GGHL06", "ZTE", "IMS_HN"),
            ("GGHL07", "ZTE", "IMS_HN"),
            ("GGHL11", "Ericsson", "IMS_HN"),
            ("GGHL12", "Ericsson", "IMS_HN"),
        ]
        cursor.executemany("INSERT INTO table_ims_routing VALUES (?, ?, ?)", default_ims)

    cursor.execute("SELECT COUNT(*) FROM table_hw_site")
    if cursor.fetchone()[0] == 0:
        default_hw = [
            ("GGPD04", "Huawei", "site"),
            ("GGPD05", "Huawei", "site"),
            ("GGPD06", "Huawei", "site"),
            ("GGPD07", "Huawei", "site"),
            ("GGHL04", "ZTE", "site"),
            ("GGHL05", "ZTE", "site"),
            ("GGHL06", "ZTE", "site"),
            ("GGHL07", "ZTE", "site"),
            ("GGHL11", "Ericsson", "site"),
            ("GGHL12", "Ericsson", "site"),
        ]
        cursor.executemany("INSERT INTO table_hw_site VALUES (?, ?, ?)", default_hw)

    conn.commit()
    conn.close()

def init_db_extended():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS table_schedules (
            table_key TEXT PRIMARY KEY,
            query TEXT,
            schedule_type TEXT,
            is_active INTEGER,
            last_run TEXT
        )
    """)
    # Insert default manual schedules if empty
    cursor.execute("SELECT COUNT(*) FROM table_schedules")
    if cursor.fetchone()[0] == 0:
        defaults = [
            ("license", "SELECT node, vendor, area, lic_bear, lic_throughput, lic_bear_uctt, lic_throughput_vhkt FROM table_license;", "manual", 0, ""),
            ("current", "SELECT node, vendor, bear_su_dung, throughput, bear_total_su_dung, bear_ims, ipv4_internet, ipv4_ims FROM table_current;", "manual", 0, ""),
            ("weight", "SELECT node, vendor, weight, new_weight, on_off FROM table_weight;", "manual", 0, ""),
            ("ims_routing", "SELECT node, vendor, ims_site FROM table_ims_routing;", "manual", 0, ""),
            ("hw_site", "SELECT node, vendor, hw_nfvi_site FROM table_hw_site;", "manual", 0, "")
        ]
        cursor.executemany("INSERT INTO table_schedules VALUES (?, ?, ?, ?, ?)", defaults)
    conn.commit()
    conn.close()

init_db()
init_db_extended()

class DBQueryRequest(BaseModel):
    connection_string: Optional[str] = None
    sql_query: str

class ExportRequest(BaseModel):
    file_name: str
    grid_data: List[Dict[str, Any]]

class SaveDataRequest(BaseModel):
    table_key: str
    rows: List[Dict[str, Any]]

@app.get("/api/get-all-data")
async def get_all_data():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Load License
    cursor.execute("SELECT * FROM table_license")
    license_rows = [dict(r) for r in cursor.fetchall()]
    
    # Load Current
    cursor.execute("SELECT * FROM table_current")
    current_rows = [dict(r) for r in cursor.fetchall()]
    
    # Load Weight
    cursor.execute("SELECT * FROM table_weight")
    weight_rows = [dict(r) for r in cursor.fetchall()]
    
    # Load IMS Routing
    cursor.execute("SELECT * FROM table_ims_routing")
    ims_rows = [dict(r) for r in cursor.fetchall()]
    
    # Load HW Site
    cursor.execute("SELECT * FROM table_hw_site")
    hw_rows = [dict(r) for r in cursor.fetchall()]
    
    # Translate DB keys back to UI template columns
    # UI License cols: ["Node", "Vendor", "License Bear", "License Throughput", "License Bear UCTT (110%)", "License Throughput VHKT"]
    license_ui = []
    for r in license_rows:
        license_ui.append({
            "Node": r["node"],
            "Vendor": r["vendor"] or "Huawei",
            "License Bear": r["lic_bear"],
            "License Throughput": r["lic_throughput"],
            "License Bear UCTT (110%)": r["lic_bear_uctt"],
            "License Throughput VHKT": r["lic_throughput_vhkt"]
        })
        
    # UI Current cols: ["Node", "Vendor", "Bear sử dụng", "Throughput", "Bear Total sử dụng", "Bear IMS", "Total IPv4 (v-internet)", "Total IPv4 (IMS)"]
    current_ui = []
    for r in current_rows:
        current_ui.append({
            "Node": r["node"],
            "Vendor": r["vendor"] or "Huawei",
            "Bear sử dụng": r["bear_su_dung"],
            "Throughput": r["throughput"],
            "Bear Total sử dụng": r["bear_total_su_dung"],
            "Bear IMS": r["bear_ims"],
            "Total IPv4 (v-internet)": r["ipv4_internet"],
            "Total IPv4 (IMS)": r["ipv4_ims"]
        })
        
    # UI Weight cols: ["Node", "Vendor", "Weight (AH)", "NEW WEIGHT (AK)", "ON/OFF"]
    weight_ui = []
    for r in weight_rows:
        weight_ui.append({
            "Node": r["node"],
            "Vendor": r["vendor"] or "Huawei",
            "Weight (AH)": r["weight"],
            "NEW WEIGHT (AK)": r["new_weight"],
            "ON/OFF": r["on_off"]
        })
        
    # UI IMS Routing cols: ["Node", "Vendor", "IMS_site"]
    ims_ui = []
    for r in ims_rows:
        ims_ui.append({
            "Node": r["node"],
            "Vendor": r["vendor"] or "Huawei",
            "IMS_site": r["ims_site"] or "IMS_HN"
        })
        
    # UI HW Site cols: ["Node", "Vendor", "HW/NFVI Site"]
    hw_ui = []
    for r in hw_rows:
        hw_ui.append({
            "Node": r["node"],
            "Vendor": r["vendor"] or "Huawei",
            "HW/NFVI Site": r["hw_nfvi_site"] or "site"
        })
        
    conn.close()
    return {
        "success": True,
        "license": license_ui,
        "current": current_ui,
        "weight": weight_ui,
        "ims_routing": ims_ui,
        "hw_site": hw_ui
    }

@app.post("/api/save-table-data")
async def save_table_data(req: SaveDataRequest):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    try:
        if req.table_key == "license":
            cursor.execute("DELETE FROM table_license")
            for r in req.rows:
                node = str(r.get("Node") or r.get("node") or "")
                if not node: continue
                cursor.execute("""
                    INSERT OR REPLACE INTO table_license (node, vendor, lic_bear, lic_throughput, lic_bear_uctt, lic_throughput_vhkt)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    node,
                    r.get("Vendor") or r.get("vendor") or "Huawei",
                    float(r.get("License Bear") or r.get("lic_bear") or 0),
                    float(r.get("License Throughput") or r.get("lic_throughput") or 0),
                    float(r.get("License Bear UCTT (110%)") or r.get("lic_bear_uctt") or 0),
                    float(r.get("License Throughput VHKT") or r.get("lic_throughput_vhkt") or 0)
                ))
        elif req.table_key == "current":
            cursor.execute("DELETE FROM table_current")
            for r in req.rows:
                node = str(r.get("Node") or r.get("node") or "")
                if not node: continue
                cursor.execute("""
                    INSERT OR REPLACE INTO table_current (node, vendor, bear_su_dung, throughput, bear_total_su_dung, bear_ims, ipv4_internet, ipv4_ims)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    node,
                    r.get("Vendor") or r.get("vendor") or "Huawei",
                    float(r.get("Bear sử dụng") or r.get("bear_su_dung") or 0),
                    float(r.get("Throughput") or r.get("throughput") or 0),
                    float(r.get("Bear Total sử dụng") or r.get("bear_total_su_dung") or 0),
                    float(r.get("Bear IMS") or r.get("bear_ims") or 0),
                    float(r.get("Total IPv4 (v-internet)") or r.get("ipv4_internet") or 0),
                    float(r.get("Total IPv4 (IMS)") or r.get("ipv4_ims") or 0)
                ))
        elif req.table_key == "weight":
            cursor.execute("DELETE FROM table_weight")
            for r in req.rows:
                node = str(r.get("Node") or r.get("node") or "")
                if not node: continue
                cursor.execute("""
                    INSERT OR REPLACE INTO table_weight (node, vendor, weight, new_weight, on_off)
                    VALUES (?, ?, ?, ?, ?)
                """, (
                    node,
                    r.get("Vendor") or r.get("vendor") or "Huawei",
                    float(r.get("Weight (AH)") or r.get("weight") or 0),
                    float(r.get("NEW WEIGHT (AK)") or r.get("new_weight") or 0),
                    int(r.get("ON/OFF") or r.get("on_off") or 1)
                ))
        elif req.table_key == "ims_routing":
            cursor.execute("DELETE FROM table_ims_routing")
            for r in req.rows:
                node = str(r.get("Node") or r.get("node") or "")
                if not node: continue
                cursor.execute("""
                    INSERT OR REPLACE INTO table_ims_routing (node, vendor, ims_site)
                    VALUES (?, ?, ?)
                """, (
                    node,
                    r.get("Vendor") or r.get("vendor") or "Huawei",
                    r.get("IMS_site") or r.get("ims_site") or "IMS_HN"
                ))
        elif req.table_key == "hw_site":
            cursor.execute("DELETE FROM table_hw_site")
            for r in req.rows:
                node = str(r.get("Node") or r.get("node") or "")
                if not node: continue
                cursor.execute("""
                    INSERT OR REPLACE INTO table_hw_site (node, vendor, hw_nfvi_site)
                    VALUES (?, ?, ?)
                """, (
                    node,
                    r.get("Vendor") or r.get("vendor") or "Huawei",
                    r.get("HW/NFVI Site") or r.get("hw_nfvi_site") or r.get("HW_Site") or "site"
                ))
        
        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/query-db")
async def query_db(req: DBQueryRequest):
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute(req.sql_query)
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
        
        result = []
        for r in rows:
            result.append(dict(zip(columns, r)))
            
        conn.close()
        return {
            "success": True,
            "columns": columns,
            "rows": result
        }
    except Exception as e:
        return {
            "success": False,
            "detail": f"Database query error: {str(e)}"
        }

@app.post("/api/export")
async def export_file(req: ExportRequest):
    # Search local download directory template first
    src_path = os.path.join("c:\\Users\\longvh3\\Downloads\\Tinh_tai_UCTT", "01072026_New_VN_Report_PS_Core_Resource_Tinhlq1.xlsx")
    if not os.path.exists(src_path):
        raise HTTPException(status_code=404, detail="Original template Excel file not found.")
            
    export_filename = f"exported_{req.file_name}"
    dest_path = os.path.join(UPLOAD_DIR, export_filename)
    shutil.copy(src_path, dest_path)
    
    try:
        wb = openpyxl.load_workbook(dest_path, data_only=False)
        ws = wb["Resource_GGSN_by_NOCPRO"]
        
        # Write modified values/formulas back
        for row in req.grid_data:
            r_idx = row.get("row_num")
            if not r_idx:
                continue
            for col_letter, cell_info in row.items():
                if col_letter in ["row_num", "Node"]:
                    continue
                if isinstance(cell_info, dict):
                    val = cell_info.get("value")
                    formula = cell_info.get("formula")
                    is_formula = cell_info.get("is_formula")
                    
                    c_idx = openpyxl.utils.column_index_from_string(col_letter)
                    
                    if is_formula and formula:
                        ws.cell(row=r_idx, column=c_idx, value=formula)
                    else:
                        try:
                            num_val = float(val)
                            ws.cell(row=r_idx, column=c_idx, value=num_val)
                        except (ValueError, TypeError):
                            ws.cell(row=r_idx, column=c_idx, value=val)
                            
        wb.save(dest_path)
        return FileResponse(
            dest_path, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=export_filename
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel export: {str(e)}")

class SaveScheduleRequest(BaseModel):
    table_key: str
    query: str
    schedule_type: str
    is_active: int

@app.get("/api/get-schedules")
async def get_schedules():
    try:
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM table_schedules")
        rows = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return {"success": True, "schedules": rows}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/save-schedule")
async def save_schedule(req: SaveScheduleRequest):
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("""
            INSERT OR REPLACE INTO table_schedules (table_key, query, schedule_type, is_active, last_run)
            VALUES (?, ?, ?, ?, COALESCE((SELECT last_run FROM table_schedules WHERE table_key = ?), ''))
        """, (req.table_key, req.query, req.schedule_type, req.is_active, req.table_key))
        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def execute_and_apply_query(table_key: str, query: str):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    try:
        cursor.execute(query)
        columns = [col[0].lower() for col in cursor.description]
        rows = cursor.fetchall()
        
        db_rows = []
        for r in rows:
            db_rows.append(dict(zip(columns, r)))
            
        if table_key == "license":
            cursor.execute("DELETE FROM table_license")
            for r in db_rows:
                node = str(r.get("node") or r.get("node_name") or "")
                if not node: continue
                cursor.execute("""
                    INSERT OR REPLACE INTO table_license (node, vendor, lic_bear, lic_throughput, lic_bear_uctt, lic_throughput_vhkt)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    node,
                    r.get("vendor") or "Huawei",
                    float(r.get("lic_bear") or r.get("license_bear") or 0),
                    float(r.get("lic_throughput") or r.get("license_throughput") or 0),
                    float(r.get("lic_bear_uctt") or r.get("license_bear_uctt_110") or 0),
                    float(r.get("lic_throughput_vhkt") or r.get("license_throughput_vhkt") or 0)
                ))
        elif table_key == "current":
            cursor.execute("DELETE FROM table_current")
            for r in db_rows:
                node = str(r.get("node") or "")
                if not node: continue
                cursor.execute("""
                    INSERT OR REPLACE INTO table_current (node, vendor, bear_su_dung, throughput, bear_total_su_dung, bear_ims, ipv4_internet, ipv4_ims)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    node,
                    r.get("vendor") or "Huawei",
                    float(r.get("bear_su_dung") or r.get("bear_sử_dụng") or 0),
                    float(r.get("throughput") or 0),
                    float(r.get("bear_total_su_dung") or r.get("bear_total_sử_dụng") or 0),
                    float(r.get("bear_ims") or 0),
                    float(r.get("ipv4_internet") or r.get("ipv4_v_internet") or r.get("total_ipv4_v_internet") or r.get("ipv4_internet_total") or 0),
                    float(r.get("ipv4_ims") or r.get("total_ipv4_ims") or 0)
                ))
        elif table_key == "weight":
            cursor.execute("DELETE FROM table_weight")
            for r in db_rows:
                node = str(r.get("node") or "")
                if not node: continue
                cursor.execute("""
                    INSERT OR REPLACE INTO table_weight (node, vendor, weight, new_weight, on_off)
                    VALUES (?, ?, ?, ?, ?)
                """, (
                    node,
                    r.get("vendor") or "Huawei",
                    float(r.get("weight") or r.get("weight_ah") or 0),
                    float(r.get("new_weight") or r.get("new_weight_ak") or 0),
                    int(r.get("on_off") or 1)
                ))
        elif table_key == "ims_routing":
            cursor.execute("DELETE FROM table_ims_routing")
            for r in db_rows:
                node = str(r.get("node") or "")
                if not node: continue
                cursor.execute("""
                    INSERT OR REPLACE INTO table_ims_routing (node, vendor, ims_site)
                    VALUES (?, ?, ?)
                """, (
                    node,
                    r.get("vendor") or "Huawei",
                    r.get("ims_site") or "IMS_HN"
                ))
        elif table_key == "hw_site":
            cursor.execute("DELETE FROM table_hw_site")
            for r in db_rows:
                node = str(r.get("node") or "")
                if not node: continue
                cursor.execute("""
                    INSERT OR REPLACE INTO table_hw_site (node, vendor, hw_nfvi_site)
                    VALUES (?, ?, ?)
                """, (
                    node,
                    r.get("vendor") or "Huawei",
                    r.get("hw_nfvi_site") or r.get("hw_site") or "site"
                ))
        conn.commit()
    except Exception as e:
        print(f"Error executing scheduled query for {table_key}: {e}")
    finally:
        conn.close()

async def schedule_runner_loop():
    await asyncio.sleep(5)
    while True:
        try:
            conn = sqlite3.connect(DB_FILE)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM table_schedules WHERE is_active = 1 AND schedule_type != 'manual'")
            active_schedules = [dict(r) for r in cursor.fetchall()]
            conn.close()
            
            now = datetime.datetime.now()
            
            for sched in active_schedules:
                table_key = sched["table_key"]
                query = sched["query"]
                schedule_type = sched["schedule_type"]
                last_run_str = sched["last_run"]
                
                should_run = False
                if not last_run_str:
                    should_run = True
                else:
                    try:
                        last_run = datetime.datetime.fromisoformat(last_run_str)
                        delta = now - last_run
                        if schedule_type == "5m" and delta.total_seconds() >= 300:
                            should_run = True
                        elif schedule_type == "30m" and delta.total_seconds() >= 1800:
                            should_run = True
                        elif schedule_type == "1h" and delta.total_seconds() >= 3600:
                            should_run = True
                        elif schedule_type == "24h" and delta.total_seconds() >= 86400:
                            should_run = True
                    except Exception:
                        should_run = True
                        
                if should_run:
                    print(f"Running scheduled query for {table_key}...")
                    execute_and_apply_query(table_key, query)
                    
                    conn = sqlite3.connect(DB_FILE)
                    cursor = conn.cursor()
                    cursor.execute("UPDATE table_schedules SET last_run = ? WHERE table_key = ?", (now.isoformat(), table_key))
                    conn.commit()
                    conn.close()
                    
            await asyncio.sleep(30)
        except asyncio.CancelledError:
            break
        except Exception as e:
            print(f"Error in schedule loop: {e}")
            await asyncio.sleep(10)

@app.on_event("startup")
async def startup_event():
    asyncio.create_task(schedule_runner_loop())

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
