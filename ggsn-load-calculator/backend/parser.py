import openpyxl
from openpyxl.utils import get_column_letter
import json
import re

def parse_excel_report(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    # 1. Parse Resource_GGSN_by_NOCPRO sheet
    ws = wb["Resource_GGSN_by_NOCPRO"]
    
    # Identify headers in Row 2
    headers = {}
    col_mapping = {}
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        val = ws.cell(row=2, column=col_idx).value
        if val:
            headers[col_letter] = str(val).strip()
            col_mapping[col_letter] = col_idx
            
    # Read GGSN node rows (Row 3 to 70 or until Node is empty)
    rows_data = []
    # Let's also read data_only values to get precalculated values
    wb_val = openpyxl.load_workbook(file_path, data_only=True)
    ws_val = wb_val["Resource_GGSN_by_NOCPRO"]
    
    for r_idx in range(3, ws.max_row + 1):
        node_name = ws.cell(row=r_idx, column=2).value
        if not node_name:
            continue
        
        row_dict = {"row_num": r_idx}
        for col_letter, header in headers.items():
            col_idx = col_mapping[col_letter]
            cell_formula = ws.cell(row=r_idx, column=col_idx).value
            cell_value = ws_val.cell(row=r_idx, column=col_idx).value
            
            is_formula = False
            formula_str = ""
            if isinstance(cell_formula, str) and cell_formula.startswith("="):
                is_formula = True
                formula_str = cell_formula
                
            row_dict[col_letter] = {
                "header": header,
                "value": cell_value,
                "formula": formula_str,
                "is_formula": is_formula
            }
        rows_data.append(row_dict)
        
    # 2. Extract reference data from other sheets for lookup
    # Sheets of interest: 'GGSN-Lic', 'Input_Lic_Rate_PS', 'Input_NOCPRO_GGSN', 'IMS', 'Input_TaiNguyenNode__GGSN_Daily'
    ref_data = {}
    
    # Let's extract 'GGSN-Lic'
    if 'GGSN-Lic' in wb.sheetnames:
        ws_lic = wb_val['GGSN-Lic']
        lic_rows = []
        for r in ws_lic.iter_rows(values_only=True):
            if any(r):
                lic_rows.append(list(r))
        ref_data['GGSN-Lic'] = lic_rows

    # 'Input_Lic_Rate_PS'
    if 'Input_Lic_Rate_PS' in wb.sheetnames:
        ws_rate = wb_val['Input_Lic_Rate_PS']
        rate_rows = []
        for r in ws_rate.iter_rows(values_only=True):
            if any(r):
                rate_rows.append(list(r))
        ref_data['Input_Lic_Rate_PS'] = rate_rows
        
    # 'Input_NOCPRO_GGSN'
    if 'Input_NOCPRO_GGSN' in wb.sheetnames:
        ws_noc = wb_val['Input_NOCPRO_GGSN']
        noc_rows = []
        for r in ws_noc.iter_rows(values_only=True):
            if any(r):
                noc_rows.append(list(r))
        ref_data['Input_NOCPRO_GGSN'] = noc_rows
        
    # 'IMS'
    if 'IMS' in wb.sheetnames:
        ws_ims = wb_val['IMS']
        ims_rows = []
        for r in ws_ims.iter_rows(values_only=True):
            if any(r):
                ims_rows.append(list(r))
        ref_data['IMS'] = ims_rows

    # Global/Static values referenced in formulas (e.g. $Z$1, $AC$1, $U$82, $T$82)
    # Z1 is =Input_TaiNguyenNode__GGSN_Daily!K2/Input_TaiNguyenNode__GGSN_Daily!G2 in data_only
    # Let's extract cells referenced in sheet Resource_GGSN_by_NOCPRO:
    # Z1, AC1, U82, T82
    global_refs = {
        "Z1": ws_val["Z1"].value if ws_val["Z1"] else None,
        "AC1": ws_val["AC1"].value if ws_val["AC1"] else None,
        "U82": ws_val["U82"].value if ws_val["U82"] else None,
        "T82": ws_val["T82"].value if ws_val["T82"] else None,
    }
    
    return {
        "headers": headers,
        "rows": rows_data,
        "ref_data": ref_data,
        "global_refs": global_refs
    }
